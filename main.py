from flask import Flask, render_template, request, send_file, redirect, url_for
import os
import csv
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import logging

logging.basicConfig(level=logging.INFO)

app = Flask(__name__)
FILENAME = os.getenv('FILENAME', 'Periodic Table of Elements.csv')
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'uploads')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --------------------- Utility Functions ---------------------
def get_parsed_formula(formula):
    pattern = r"([A-Z][a-z]*)(\d*\.?\d*)"
    elements = re.findall(pattern, formula)
    ratios = {}
    for element, coefficient in elements:
        ratio = float(coefficient) if coefficient else 1
        ratios[element] = ratio
    return ratios

def read_molar_masses(filename):
    molar_masses = {}
    with open(filename, mode='r', newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            element, molar_mass = row
            molar_masses[element.strip()] = float(molar_mass.strip())
    return molar_masses

def calculate_masses(ratios, total_mass, molar_masses):
    total_molar_mass = sum(ratios[element] * molar_masses[element] for element in ratios)
    masses = {}
    for element in ratios:
        element_mass_fraction = (ratios[element] * molar_masses[element]) / total_molar_mass
        masses[element] = element_mass_fraction * total_mass
    logging.info("Masses calculated: %s", masses)
    return masses

def calculate_masses_with_known_element(ratios, known_element, known_mass, molar_masses):
    total_molar_mass = sum(ratios[element] * molar_masses[element] for element in ratios)
    known_element_molar_mass = ratios[known_element] * molar_masses[known_element]
    scaling_factor = known_mass / known_element_molar_mass
    masses = {}
    for element in ratios:
        masses[element] = scaling_factor * ratios[element] * molar_masses[element]
    logging.info("Masses calculated: %s", masses)
    return masses

def process_formulas(formulas, total_masses, molar_masses):
    wb = Workbook()
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for total_mass in total_masses:
        ws = wb.create_sheet(title=f"{total_mass:.2f}")
        max_elements = max(len(get_parsed_formula(formula)) for formula in formulas)
        headers = ['Formula', 'Total Mass']
        for i in range(max_elements):
            headers.extend([f'Element{i+1}', f'Mass{i+1}'])
        ws.append(headers)

        for formula in formulas:
            ratios = get_parsed_formula(formula)
            masses = calculate_masses(ratios, total_mass, molar_masses)
            logging.info("Masses calculated: %s", masses)
            row = [formula, total_mass]
            for element, mass in masses.items():
                row.extend([element, f"{mass:.4f}"])
            ws.append(row)

        for col in range(4, 4 + 2 * max_elements, 2):
            for row in range(2, len(formulas) + 2):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.remove(wb['Sheet'])
    output_path = os.path.join(UPLOAD_FOLDER, 'calculated_output.xlsx')
    wb.save(output_path)
    return output_path

def read_formulas_from_txt_file(filepath):
    with open(filepath, 'r') as file:
        formulas = [line.strip() for line in file.readlines()]
    return formulas

def read_formulas_from_excel(filepath):
    df = pd.read_excel(filepath, header=None)
    return df[0].tolist()

# --------------------- Flask Routes ---------------------
@app.route('/', methods=['GET', 'POST'])
def index():
    error = None
    if request.method == 'POST':
        try:
            mode = request.form.get('mode')
            molar_masses = read_molar_masses(FILENAME)

            if mode == 'single':
                formula = request.form['formula']
                if not re.match(r"^[A-Za-z0-9\s]+$", formula):
                    raise ValueError("Invalid formula format.")
                total_mass = float(request.form['total_mass'])
                ratios = get_parsed_formula(formula)
                wrong_elements = [e for e in ratios if e not in molar_masses]
                if wrong_elements:
                    raise ValueError(f"Invalid element(s): {', '.join(wrong_elements)}")
                masses = calculate_masses(ratios, total_mass, molar_masses)
                logging.info("Masses calculated: %s", masses)
                return render_template('results.html', formula=formula, total_mass=total_mass, masses=masses)

            elif mode == 'batch':
                formulas_text = request.form['formulas']
                formulas = [line.strip() for line in formulas_text.strip().split('\n') if line.strip()]
                valid_formulas = [f for f in formulas if all(el in molar_masses for el in get_parsed_formula(f))]
                total_masses = [0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50]
                output_path = process_formulas(valid_formulas, total_masses, molar_masses)
                return send_file(output_path, as_attachment=True)

            elif mode == 'known':
                formula = request.form['formula_known']
                if not re.match(r"^[A-Za-z0-9\s]+$", formula):
                    raise ValueError("Invalid formula format.")
                known_element = request.form['known_element']
                known_mass = float(request.form['known_mass'])
                ratios = get_parsed_formula(formula)
                if known_element not in ratios:
                    raise ValueError(f"'{known_element}' not in formula")
                masses = calculate_masses_with_known_element(ratios, known_element, known_mass, molar_masses)
                logging.info("Masses calculated: %s", masses)
                return render_template('results_known.html', formula=formula, known_element=known_element, known_mass=known_mass, masses=masses)

        except Exception as e:
            logging.error("An error occurred: %s", str(e))
            error = str(e)
            return render_template('index.html', error="Invalid formula. Please check your input.")

    return render_template('index.html', error=error)

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        # Check if a file is uploaded
        if 'file' not in request.files:
            raise ValueError("No file uploaded.")
        
        file = request.files['file']
        if file.filename == '':
            raise ValueError("No file selected.")
        
        # Save the uploaded file
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)
        logging.info("File uploaded: %s", file_path)

        # Process the file based on its extension
        if file.filename.endswith('.txt'):
            formulas = read_formulas_from_txt_file(file_path)
        elif file.filename.endswith('.xlsx'):
            formulas = read_formulas_from_excel(file_path)
        else:
            raise ValueError("Unsupported file type. Please upload a .txt or .xlsx file.")

        # Validate formulas
        molar_masses = read_molar_masses(FILENAME)
        valid_formulas = [f for f in formulas if all(el in molar_masses for el in get_parsed_formula(f))]
        if not valid_formulas:
            raise ValueError("No valid formulas found in the file.")

        # Process formulas and generate Excel output
        total_masses = [0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50]
        output_path = process_formulas(valid_formulas, total_masses, molar_masses)

        # Return the generated Excel file
        return send_file(output_path, as_attachment=True)

    except Exception as e:
        logging.error("An error occurred: %s", str(e))
        return render_template('index.html', error=str(e))
    
@app.route('/fast_oxidizing', methods=['POST'])
def fast_oxidizing():
    try:
        # Get form data
        known_element = request.form['known_element']
        formula = request.form['formula']
        known_mass = float(request.form['known_mass'])

        # Validate the formula
        molar_masses = read_molar_masses(FILENAME)
        ratios = get_parsed_formula(formula)
        wrong_elements = [e for e in ratios if e not in molar_masses]
        if wrong_elements:
            raise ValueError(f"Invalid element(s): {', '.join(wrong_elements)}")
        if known_element not in ratios:
            raise ValueError(f"The element '{known_element}' is not present in the formula.")

        # Calculate masses
        masses = calculate_masses_with_known_element(ratios, known_element, known_mass, molar_masses)
        logging.info("Masses calculated: %s", masses)

        # Render results
        return render_template('results.html', formula=formula, known_element=known_element, known_mass=known_mass, masses=masses)

    except Exception as e:
        logging.error("An error occurred: %s", str(e))
        return render_template('index.html', error=str(e))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))  # Use the PORT environment variable or default to 5000
    app.run(host='0.0.0.0', port=port)