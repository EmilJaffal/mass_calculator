import click
import os
import csv
import util.batch_processing as bp
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def calculate_masses(ratios, total_mass, molar_masses):
    total_molar_mass = sum(ratios[element] * molar_masses[element] for element in ratios)
    
    masses = {}
    for element in ratios:
        element_mass_fraction = (ratios[element] * molar_masses[element]) / total_molar_mass
        masses[element] = element_mass_fraction * total_mass
    return masses

def calculate_masses_with_known_element(ratios, known_element, known_mass, molar_masses):
    total_molar_mass = sum(ratios[element] * molar_masses[element] for element in ratios)
    known_element_molar_mass = ratios[known_element] * molar_masses[known_element]
    scaling_factor = known_mass / known_element_molar_mass
    
    masses = {}
    for element in ratios:
        masses[element] = scaling_factor * ratios[element] * molar_masses[element]
    
    return masses

def process_formulas(formulas, filename, total_masses=[0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50]):
    molar_masses = bp.read_molar_masses(filename)
    wb = Workbook()
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for total_mass in total_masses:
        ws = wb.create_sheet(title=f"{total_mass:.2f}")
        max_elements = max(len(bp.get_parsed_formula(formula)) for formula in formulas)
        headers = ['Formula', 'Total Mass']
        for i in range(max_elements):
            headers.extend([f'Element{i+1}', f'Mass{i+1}'])
        ws.append(headers)
        
        for formula in formulas:
            ratios = bp.get_parsed_formula(formula)
            masses = calculate_masses(ratios, total_mass, molar_masses)
            row = [formula, total_mass]
            for element, mass in masses.items():
                row.extend([element, f"{mass:.4f}"])
            ws.append(row)
        
        for col in range(4, 4 + 2 * max_elements, 2):
            for row in range(2, len(formulas) + 2):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.remove(wb['Sheet'])
    base_filename = 'calculated'
    extension = '.xlsx'
    n = 1
    while os.path.exists(f"{base_filename}_{n}{extension}"):
        n += 1
    unique_filename = f"{base_filename}_{n}{extension}"
    
    wb.save(unique_filename)
    print(f"File saved as {unique_filename}")

def read_formulas_from_text():
    formulas = []
    print("Paste the formulas (one per line), and press 'Enter' on an empty line when finished:")
    while True:
        line = input()
        if line == '':
            break
        formulas.append(line.strip())
    return formulas

def read_formulas_from_txt_file(filepath):
    with open(filepath, 'r') as file:
        formulas = [line.strip() for line in file.readlines()]
    return formulas

def read_formulas_from_excel(filepath):
    df = pd.read_excel(filepath, header=None)
    return df[0].tolist()

def read_molar_masses(filename):
    molar_masses = {}
    with open(filename, mode='r', newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            element, molar_mass = row
            molar_masses[element.strip()] = float(molar_mass.strip())
    return molar_masses

def get_parsed_formula(formula):
    pattern = r"([A-Z][a-z]*)(\d*\.?\d*)"
    elements = re.findall(pattern, formula)
    ratios = {}
    for element, coefficient in elements:
        ratio = float(coefficient) if coefficient else 1
        ratios[element] = ratio
    return ratios

def list_files_by_extension(extension, exclude=[]):
    current_directory = os.path.dirname(os.path.abspath(__file__))
    return [f for f in os.listdir(current_directory) if f.endswith(extension) and f not in exclude]

def finish_calculation():
    bold_text = "\033[1mDo you want to finish?\033[0m"
    return click.confirm(bold_text, default=True)

@click.command()
@click.option('--filename', default='Periodic Table of Elements.csv', help='CSV file containing element molar masses.')
def main(filename):
    while True:
        click.echo("Choose an input method:")
        click.echo("1. Enter the chemical formula")
        click.echo("2. Paste column of text with formulas")
        click.echo("3. Read from txt file with formulas")
        click.echo("4. Read from Excel file with formulas")
        click.echo("5. Calculate masses based on the known mass of one element")

        choice = click.prompt("Enter your choice", type=int)

        molar_masses = read_molar_masses(filename)

        if choice == 1:
            formula = click.prompt("Enter the chemical formula of the mixture (e.g., CdCu4Ho, DyCo2)", type=str)
            try:
                ratios = get_parsed_formula(formula)
                wrong_elements = [element for element in ratios if element not in molar_masses]

                if wrong_elements:
                    click.echo(f"The formula you entered is wrong. Wrong elements: {', '.join(wrong_elements)}")
                else:
                    total_mass = click.prompt("Enter the total mass of the mixture in grams", type=float)
                    masses = calculate_masses(ratios, total_mass, molar_masses)
                    click.echo("Element masses:")
                    for element, mass in masses.items():
                        click.echo(f"{element}: {mass:.4f} g")
                    click.echo(f"Total mass: {total_mass} g")
            except Exception as e:
                click.echo(f"An error occurred: {e}")

        elif choice == 2:
            formulas = read_formulas_from_text()
            valid_formulas = []

            for formula in formulas:
                try:
                    ratios = get_parsed_formula(formula)
                    wrong_elements = [element for element in ratios if element not in molar_masses]

                    if wrong_elements:
                        click.echo(f"The formula '{formula}' is wrong. Wrong elements: {', '.join(wrong_elements)}")
                    else:
                        valid_formulas.append(formula)
                except Exception as e:
                    click.echo(f"An error occurred while processing '{formula}': {e}")

            if valid_formulas:
                process_formulas(valid_formulas, filename)

        elif choice == 3:
            txt_files = list_files_by_extension('.txt', exclude=['README.txt'])
            if not txt_files:
                click.echo("No .txt files found in the current directory.")
                continue
            
            click.echo("Available .txt files:")
            for i, file in enumerate(txt_files, 1):
                click.echo(f"{i}. {file}")

            file_choice = click.prompt("Choose a .txt file by number", type=int)
            if file_choice < 1 or file_choice > len(txt_files):
                click.echo("Invalid choice.")
                continue
            
            filepath = txt_files[file_choice - 1]
            formulas = read_formulas_from_txt_file(filepath)
            valid_formulas = []

            for formula in formulas:
                try:
                    ratios = get_parsed_formula(formula)
                    wrong_elements = [element for element in ratios if element not in molar_masses]

                    if wrong_elements:
                        click.echo(f"The formula '{formula}' is wrong. Wrong elements: {', '.join(wrong_elements)}")
                    else:
                        valid_formulas.append(formula)
                except Exception as e:
                    click.echo(f"An error occurred while processing '{formula}': {e}")

            if valid_formulas:
                process_formulas(valid_formulas, filename)

        elif choice == 4:
            xlsx_files = list_files_by_extension('.xlsx', exclude=[f"calculated_{n}.xlsx" for n in range(1, 1001)])
            if not xlsx_files:
                click.echo("No .xlsx files found in the current directory.")
                continue
            
            click.echo("Available .xlsx files:")
            for i, file in enumerate(xlsx_files, 1):
                click.echo(f"{i}. {file}")

            file_choice = click.prompt("Choose an .xlsx file by number", type=int)
            if file_choice < 1 or file_choice > len(xlsx_files):
                click.echo("Invalid choice.")
                continue
            
            filepath = xlsx_files[file_choice - 1]
            formulas = read_formulas_from_excel(filepath)
            valid_formulas = []

            for formula in formulas:
                try:
                    ratios = get_parsed_formula(formula)
                    wrong_elements = [element for element in ratios if element not in molar_masses]

                    if wrong_elements:
                        click.echo(f"The formula '{formula}' is wrong. Wrong elements: {', '.join(wrong_elements)}")
                    else:
                        valid_formulas.append(formula)
                except Exception as e:
                    click.echo(f"An error occurred while processing '{formula}': {e}")

            if valid_formulas:
                process_formulas(valid_formulas, filename)

        
        elif choice == 5:
            while True:
                known_element = click.prompt("Enter the element with known mass (e.g., Os)", type=str)
                
                while True:
                    formula = click.prompt("Enter the chemical formula of the mixture (e.g., GdOsIn)", type=str)
                    known_mass = click.prompt(f"Enter the known mass of {known_element} in grams", type=float)
    
                    try:
                        ratios = get_parsed_formula(formula)
                        wrong_elements = [element for element in ratios if element not in molar_masses]
    
                        if wrong_elements:
                            click.echo(f"The formula you entered is wrong. Wrong elements: {', '.join(wrong_elements)}")
                        elif known_element not in ratios:
                            click.echo(f"The element '{known_element}' is not present in the formula.")
                        else:
                            masses = calculate_masses_with_known_element(
                                ratios, known_element, known_mass, molar_masses
                            )
                            click.echo("Element masses:")
                            for element, mass in masses.items():
                                click.echo(f"{element}: {mass:.4f} g")
                            
                    except Exception as e:
                        click.echo(f"An error occurred: {e}")
    
                    if click.confirm("Do you want to change the element or finish?", default=True):
                        break
                        
                if click.confirm("Finish?", default=True):
                    break

        else:
            click.echo("Invalid choice. Please restart the program and choose a valid option.")
            continue 
        
        if finish_calculation():
            break

if __name__ == '__main__':
    main()